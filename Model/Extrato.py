from datetime import datetime
from openpyxl.reader.excel import load_workbook


def _procurar_trasacoes(leitura):
    inicio = 0
    for linha in leitura[f"A1:N{leitura.max_row}"]:
        if linha[0].value == "Transações":
            inicio = linha[0].row + 2
        if inicio > 0:
            if linha[0].value is None:
                lista = list(leitura[f"A{inicio}:N{linha[0].row - 1}"])
                novalista = []
                for i in range(len(lista)):
                    novalinha = []
                    for j in range(len(lista[i])):
                        if lista[i][j].value is None:
                            novalinha.append('')
                        elif j == 0:
                            # print(lista[i][j].value[2:16].replace(".", "/"))

                            novalinha.append(datetime.strptime(lista[i][j].value.replace(".", "/"), '%Y/%m/%d %H:%M:%S'))
                        else:
                            novalinha.append(str(lista[i][j].value))
                    novalista.append(novalinha)
                return novalista


class Extrato:
    def __init__(self, filename):
        print("vai ler extrato agr")
        leitura = load_workbook(f"./folder/{filename}")
        leitura = leitura.active
        print("leu agr")

        self.nome = filename
        self.tab = _procurar_trasacoes(leitura)
        self.cliente = leitura["D2"].value
        self.conta = int(leitura["D3"].value[:6])
        self.max_row = leitura.max_row

    def filtrar_tab_por(self, ultima_linha_dados):
        if ultima_linha_dados == "Cadastrar":
            print(ultima_linha_dados)
        if ultima_linha_dados != "Cadastrar":
            teste = 'false'
            for i in range(len(self.tab)):
                # 6
                # data2 = datetime.strptime(ultima_linha_dados[5].replace("-", "/"), '%d/%m/%y %H:%M')
                if self.tab[i][1] == ultima_linha_dados[6]:
                    teste = 'true'
                    print('entrou no if')
                    try:
                        print("Acabou")
                        print(self.tab[i])
                        self.tab[i + 1]
                    except IndexError:
                        self.tab = None
                        raise ValueError(f"Extrato já cadastrado!")
                    else:
                        numero = i + 1
                        self.tab[:numero] = []
                        break

        if teste == 'false':
            raise ValueError(f"Erro de continuidade! Não foi possivel encontrar a ultima linha da planilha no extrato.")

        print("Acabou")